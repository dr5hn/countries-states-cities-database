#!/usr/bin/env python3
"""Luxembourg -> contributions/postcodes/LU.json importer for issue #1039.

Source data
-----------
The official ``CACLR`` registry (Centre des Adresses du Cadastre du
Luxembourg / Registre national des localités et des rues), published
under CC-Zero by the Luxembourgish government on data.public.lu, is
the canonical address reference.

The xlsx contains a denormalised join sheet ``TR.DiCaCoLo.RuCp`` with
columns:
    DISTRICT_NOM, CANTON_NOM, COMMUNE_NOM, LOCALITE_NOM, RUE_NOM, CODE_POSTAL

Source URL: https://download.data.public.lu/resources/registre-national-des-localites-et-des-rues/.../caclr.xlsx

What this script does
---------------------
1. Resolves the latest ``caclr.xlsx`` URL via the data.public.lu API
   (URL is date-stamped and rotates on every refresh).
2. Fetches the xlsx via urllib (curl is blocked).
3. Parses ``TR.DiCaCoLo.RuCp`` with openpyxl, deduplicates to unique
   ``(code, locality, canton)`` tuples.
4. Maps the 13 source canton labels (12 cantons + the
   ``LUXEMBOURG-VILLE`` capital-city sub-classification) to CSC's
   12 iso2 codes via SOURCE_TO_ISO2.
5. Skips 118 records with `?` postcode (new streets without
   assigned codes).
6. Writes contributions/postcodes/LU.json idempotently.

Why xlsx (not the population CSV)
---------------------------------
The simpler ``rnpp-code-postal.csv`` ships only postcode + population
counts — no canton FK, no locality name. Only ``caclr.xlsx`` carries
the canton/commune/locality joins required for full state FK
resolution.

License
-------
CC-Zero (public domain). No attribution required, but each row carries
``source: "caclr-data-public-lu"`` for export-time provenance.

Usage
-----
    python3 bin/scripts/sync/import_luxembourg_postcodes.py
"""

from __future__ import annotations

import argparse
import io
import json
import re
import sys
import urllib.request
from pathlib import Path
from typing import Dict, List

import openpyxl


DATASET_API_URL = (
    "https://data.public.lu/api/1/datasets/"
    "registre-national-des-localites-et-des-rues/"
)

# Source CANTON_NOM (uppercase) -> CSC iso2.
SOURCE_TO_ISO2: Dict[str, str] = {
    "CAPELLEN": "CA",
    "CLERVAUX": "CL",
    "DIEKIRCH": "DI",
    "ECHTERNACH": "EC",
    "ESCH-SUR-ALZETTE": "ES",
    "GREVENMACHER": "G",
    "LUXEMBOURG": "L",
    "LUXEMBOURG-VILLE": "L",  # capital-city administrative sub-entity
    "MERSCH": "ME",
    "REDANGE": "RD",
    "REMICH": "RM",
    "VIANDEN": "VD",
    "WILTZ": "WI",
}


def resolve_xlsx_url() -> str:
    req = urllib.request.Request(
        DATASET_API_URL, headers={"User-Agent": "csc-database-postcode-importer"}
    )
    with urllib.request.urlopen(req, timeout=20) as r:
        meta = json.loads(r.read())
    for res in meta.get("resources", []):
        if res.get("format") == "xlsx" and "caclr" in (res.get("title") or "").lower():
            return res["url"]
    raise RuntimeError("caclr.xlsx not found in dataset resources")


def fetch_bytes(url: str) -> bytes:
    req = urllib.request.Request(
        url, headers={"User-Agent": "csc-database-postcode-importer"}
    )
    with urllib.request.urlopen(req, timeout=120) as r:
        return r.read()


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", default=None, help="local xlsx (skip fetch)")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    if args.input:
        raw = Path(args.input).read_bytes()
    else:
        url = resolve_xlsx_url()
        print(f"Fetching {url}")
        raw = fetch_bytes(url)
    print(f"xlsx size: {len(raw):,} bytes")

    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    if "TR.DiCaCoLo.RuCp" not in wb.sheetnames:
        print("ERROR: expected sheet 'TR.DiCaCoLo.RuCp' missing", file=sys.stderr)
        return 2
    sh = wb["TR.DiCaCoLo.RuCp"]

    project_root = Path(__file__).resolve().parents[3]
    countries = json.load(
        (project_root / "contributions/countries/countries.json").open(encoding="utf-8")
    )
    lu_country = next((c for c in countries if c.get("iso2") == "LU"), None)
    if lu_country is None:
        print("ERROR: LU not in countries.json", file=sys.stderr)
        return 2
    regex = re.compile(lu_country.get("postal_code_regex") or ".*")

    states = json.load(
        (project_root / "contributions/states/states.json").open(encoding="utf-8")
    )
    lu_states = [s for s in states if s.get("country_id") == lu_country["id"]]
    state_by_iso2: Dict[str, dict] = {
        s["iso2"]: s for s in lu_states if s.get("iso2")
    }
    print(
        f"Country: Luxembourg (id={lu_country['id']}); "
        f"states indexed: {len(lu_states)}"
    )

    seen: set = set()
    records: List[dict] = []
    skipped_no_code = 0
    skipped_unknown_code = 0
    skipped_bad_regex = 0
    skipped_no_state = 0
    matched_state = 0
    unknown_canton: Dict[str, int] = {}
    iter_rows = sh.iter_rows(values_only=True)
    next(iter_rows)  # header

    for row in iter_rows:
        district, canton, commune, locality, street, code = row
        if not code:
            skipped_no_code += 1
            continue
        code_str = str(code).strip()
        if code_str == "?":
            skipped_unknown_code += 1
            continue
        # The xlsx writes codes as numbers; re-pad to 4 digits.
        if code_str.isdigit():
            code_str = code_str.zfill(4)
        if not regex.match(code_str):
            skipped_bad_regex += 1
            continue

        canton_label = (canton or "").strip()
        locality_str = (locality or "").strip()
        commune_str = (commune or "").strip()

        iso2 = SOURCE_TO_ISO2.get(canton_label)
        state = state_by_iso2.get(iso2) if iso2 else None
        if state is None:
            unknown_canton[canton_label] = unknown_canton.get(canton_label, 0) + 1

        # Locality preference: locality_nom (canonical settlement)
        loc_for_key = locality_str or commune_str
        key = (code_str, loc_for_key.lower(), canton_label.lower())
        if key in seen:
            continue
        seen.add(key)

        record: Dict[str, object] = {
            "code": code_str,
            "country_id": int(lu_country["id"]),
            "country_code": "LU",
        }
        if state is not None:
            record["state_id"] = int(state["id"])
            record["state_code"] = state.get("iso2")
            matched_state += 1
        else:
            skipped_no_state += 1
        if loc_for_key:
            record["locality_name"] = loc_for_key
        record["type"] = "full"
        record["source"] = "caclr-data-public-lu"
        records.append(record)

    print(f"Skipped (no code):       {skipped_no_code:,}")
    print(f"Skipped ('?' code):      {skipped_unknown_code:,}")
    print(f"Skipped (regex fail):    {skipped_bad_regex:,}")
    print(f"Skipped (no state FK):   {skipped_no_state:,}")
    print(f"Records emitted:         {len(records):,}")
    pct = matched_state * 100 // max(1, len(records))
    print(f"  with state:            {matched_state:,} ({pct}%)")
    if unknown_canton:
        print("Unknown canton labels (not in SOURCE_TO_ISO2):")
        for c, n in sorted(unknown_canton.items(), key=lambda x: -x[1]):
            print(f"  {c!r}: {n}")

    if args.dry_run:
        return 0

    target = project_root / "contributions/postcodes/LU.json"
    target.parent.mkdir(parents=True, exist_ok=True)
    if target.exists():
        with target.open(encoding="utf-8") as f:
            existing = json.load(f)
        existing_seen = {
            (r["code"], (r.get("locality_name") or "").lower()) for r in existing
        }
        merged = list(existing)
        for r in records:
            key = (r["code"], (r.get("locality_name") or "").lower())
            if key not in existing_seen:
                merged.append(r)
                existing_seen.add(key)
        merged.sort(key=lambda r: (r["code"], r.get("locality_name", "")))
    else:
        merged = sorted(records, key=lambda r: (r["code"], r.get("locality_name", "")))

    with target.open("w", encoding="utf-8") as f:
        json.dump(merged, f, ensure_ascii=False, indent=2)
        f.write("\n")
    size_kb = target.stat().st_size / 1024
    print(
        f"\n[OK] Wrote {target.relative_to(project_root)} "
        f"({len(merged):,} rows, {size_kb:.0f} KB)"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
