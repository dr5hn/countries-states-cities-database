#!/usr/bin/env python3
"""Brazil CEP -> contributions/postcodes/BR.json importer for issue #1039.

Source data
-----------
The community-maintained ``Maahzuka/database-CEPS`` archive (despite the
"CEPS" name suggesting it spans Latin America, the data inside is purely
Brazilian Correios CEP coverage) redistributes a 6,615-row Excel file
covering every Brazilian municipality with its CEP range:

  https://github.com/Maahzuka/database-CEPS

Each row represents a municipality with:
  ID | UF (Brazilian 2-letter state code) | REGIAO | LOCALIDADE |
  LOCALIDADE_SEM_ACENTOS | FAIXA_DE_CEP (range) | CEP_INICIAL | CEP_FINAL
  | SITUACAO | TIPO_DE_FAIXA | LATITUDE (comma decimal) | LONGITUDE
  | COD_GEOGRAFICO_* | COD_IBGE | MICRORREGIAO | MESORREGIAO | CATEGORIA
  | ALTITUDE | LOCALIZACAO | LOCALIZACAO_SEM_ACENTOS

What this script does
---------------------
1. Reads the .xlsx via openpyxl (read-only mode, streamed)
2. Picks ONE record per municipality using its CEP_INICIAL as the
   canonical postcode. Format converted to "#####-###" to match regex.
3. Resolves state via UF -> states.iso2 direct match (BR's 27 states
   use the same 2-letter codes in both source and CSC dataset)
4. Coordinates come from the source (lat/lng with comma decimal —
   converted to dot decimal)
5. Writes contributions/postcodes/BR.json (~6,600 records)

Note about coverage
-------------------
This is municipality-level coverage, not individual-CEP. Brazil has
~145,000 unique CEPs total but Correios paywalls the full address-level
data. The municipality CEP range provides one canonical "primary" code
per place — useful for geocoding and form auto-fill while keeping the
PR scope tight.

License & attribution
---------------------
- Mirror: github.com/Maahzuka/database-CEPS (community redistribution)
- Original publisher: Correios (Brazilian postal service)
- Each row records source: "correios"

Usage
-----
    pip install openpyxl
    python3 -c "import urllib.request; urllib.request.urlretrieve(
      'https://raw.githubusercontent.com/Maahzuka/database-CEPS/main/ceps.xlsx',
      '/tmp/mx_ceps.xlsx')"

    python3 bin/scripts/sync/import_brazil_postcodes.py
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Dict, List, Optional

try:
    import openpyxl  # type: ignore
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    raise SystemExit(2)


def parse_coord(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip().replace(",", ".")  # comma -> dot decimal
    if not s:
        return None
    try:
        f = float(s)
        if abs(f) > 180:
            return None
        return f"{f:.8f}".rstrip("0").rstrip(".") or "0"
    except ValueError:
        return None


def format_cep(raw) -> Optional[str]:
    """Convert 8-digit CEP to '#####-###' form expected by regex."""
    if raw is None:
        return None
    s = str(raw).strip().replace("-", "")
    if not s.isdigit() or len(s) != 8:
        return None
    return f"{s[:5]}-{s[5:]}"


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", default="/tmp/mx_ceps.xlsx",
                        help="Path to ceps.xlsx (default: /tmp/mx_ceps.xlsx)")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    src = Path(args.input)
    if not src.exists():
        print(f"ERROR: input not found: {src}", file=sys.stderr)
        return 2

    project_root = Path(__file__).resolve().parents[3]
    countries = json.load((project_root / "contributions/countries/countries.json").open(encoding="utf-8"))
    br = next((c for c in countries if c.get("iso2") == "BR"), None)
    if br is None:
        print("ERROR: BR not in countries.json", file=sys.stderr)
        return 2
    states = json.load((project_root / "contributions/states/states.json").open(encoding="utf-8"))
    br_states = [s for s in states if s.get("country_id") == br["id"]]
    state_by_iso2: Dict[str, dict] = {(s.get("iso2") or "").upper(): s for s in br_states if s.get("iso2")}
    print(f"Country: Brazil (id={br['id']}); states indexed: {len(state_by_iso2)}")

    wb = openpyxl.load_workbook(src, read_only=True)
    ws = wb.active
    print(f"Sheet: {ws.title}, rows: {ws.max_row}")

    # Read header
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    col_idx = {name: i for i, name in enumerate(header)}

    by_code: Dict[str, dict] = {}
    bad = 0
    for row in rows:
        cep_raw = row[col_idx.get("CEP_INICIAL", -1)] if col_idx.get("CEP_INICIAL", -1) >= 0 else None
        code = format_cep(cep_raw)
        if not code:
            bad += 1
            continue
        if code in by_code:
            continue
        by_code[code] = {
            "uf": (row[col_idx["UF"]] or "").strip().upper() if col_idx.get("UF") is not None else "",
            "locality": (row[col_idx["LOCALIDADE"]] or "").strip() if col_idx.get("LOCALIDADE") is not None else "",
            "lat": row[col_idx.get("LATITUDE", -1)] if col_idx.get("LATITUDE", -1) >= 0 else None,
            "lng": row[col_idx.get("LONGITUDE", -1)] if col_idx.get("LONGITUDE", -1) >= 0 else None,
        }

    print(f"Skipped malformed:   {bad:,}")
    print(f"Unique CEPs:         {len(by_code):,}")

    records: List[dict] = []
    matched_state = 0
    matched_coord = 0
    for code in sorted(by_code):
        meta = by_code[code]
        record = {
            "code": code,
            "country_id": int(br["id"]),
            "country_code": "BR",
        }
        state = state_by_iso2.get(meta["uf"])
        if state is not None:
            record["state_id"] = int(state["id"])
            record["state_code"] = meta["uf"]
            matched_state += 1
        if meta["locality"]:
            record["locality_name"] = meta["locality"]
        record["type"] = "full"
        lat = parse_coord(meta["lat"])
        lng = parse_coord(meta["lng"])
        if lat is not None and lng is not None:
            record["latitude"] = lat
            record["longitude"] = lng
            matched_coord += 1
        record["source"] = "correios"
        records.append(record)

    print(f"Records:        {len(records):,}")
    print(f"  with state:   {matched_state:,} ({matched_state*100//max(1,len(records))}%)")
    print(f"  with coords:  {matched_coord:,} ({matched_coord*100//max(1,len(records))}%)")

    if args.dry_run:
        return 0

    target = project_root / "contributions/postcodes/BR.json"
    if target.exists():
        with target.open(encoding="utf-8") as f:
            existing = json.load(f)
        seen = {(r["code"], (r.get("locality_name") or "").lower()) for r in existing}
        merged = list(existing)
        for r in records:
            key = (r["code"], (r.get("locality_name") or "").lower())
            if key not in seen:
                merged.append(r)
                seen.add(key)
        merged.sort(key=lambda r: (r["code"], r.get("locality_name", "")))
    else:
        merged = sorted(records, key=lambda r: (r["code"], r.get("locality_name", "")))

    with target.open("w", encoding="utf-8") as f:
        json.dump(merged, f, ensure_ascii=False, indent=2)
        f.write("\n")
    size_mb = target.stat().st_size / (1024 * 1024)
    print(f"\n[OK] Wrote {target.relative_to(project_root)} ({len(merged):,} rows, {size_mb:.1f} MB)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
